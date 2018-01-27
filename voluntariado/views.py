# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from voluntariado.models import Voluntario
from django.db.models import Q
from django.http.response import HttpResponseRedirect
from voluntariado.forms import VoluntarioForm, ConsultaForm, ReferenciaForm
from django.urls.base import reverse
from django.core.paginator import Paginator
from datetime import datetime
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView

# Create your views here.
@login_required
def home (request):
    templates = 'voluntariado/voluntario_master.html'
    return render(request, templates)

@login_required
def voluntarios(request):
    
    
    
    templates = 'voluntariado/voluntarios.html'
    voluntarios = Voluntario.objects.filter(estado = Voluntario.ESTADO_ACTIVO)
    query = ''
    if 'query' in request.GET:
        query = request.GET['query']
        voluntarios = voluntarios.filter(nombres__istartswith = query)
    
    paginator = Paginator(voluntarios, 5)
    page = 1
    
    if 'page' in request.GET:
        page = int(request.GET['page'])

    pagina = paginator.page(page)


    data = {
            
            'voluntarios': pagina,
            'query': query,
            
        }
    return render(request, templates, data)

@login_required
def voluntario(request, voluntario_id = None):
    template = 'voluntariado/voluntario_crear.html'
    form = VoluntarioForm()
    voluntario = None
    
    if voluntario_id is not None:
        voluntario = Voluntario.objects.get(id = voluntario_id)
        form = VoluntarioForm(instance = voluntario)
            
    if request.method == 'POST':
        form = VoluntarioForm(request.POST, request.FILES, instance = voluntario)
        (form.data)
    if form.is_valid():
        voluntario = form.save()
        return HttpResponseRedirect(reverse('voluntarios'))
        
            #request.session['msg'] = "Empleado creado con �xito"
        
    data = {
            'form':form,
            
        }
    return render(request, template, data)

@login_required
def update(request, voluntario_id ):
    voluntario = Voluntario.objects.get(id = voluntario_id )
    voluntario.estado = "1"
    voluntario.save() 
    return HttpResponseRedirect(reverse('voluntarios'))




class ReportePersonasExcel(TemplateView):
     
    #Usamos el m�todo get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        voluntarios = Voluntario.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE VOLUNTARIOS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'ID'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'APELLIDO '
        ws['E3'] = 'CORREO'       
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for voluntario in voluntarios:
            ws.cell(row=cont,column=2).value =  voluntario.id
            ws.cell(row=cont,column=3).value = voluntario.nombres
            ws.cell(row=cont,column=4).value = voluntario.apellidos
            ws.cell(row=cont,column=5).value = voluntario.correo
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonas.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


@login_required
def referencia(request):
    template = 'referencia/referencia.html'
    form = ReferenciaForm()
    referencia = None
    if request.method == 'POST':
        form = ReferenciaForm(request.POST, request.FILES, instance = referencia)
        (form.data)
    if form.is_valid():
        referencia = form.save()
        return HttpResponseRedirect(reverse('voluntario'))

    data = {
            'form':form,
            
        }
    return render(request, template, data)
